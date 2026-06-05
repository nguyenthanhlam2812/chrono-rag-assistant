"""Run ChronoRAG chatbot regression checks from an Excel test-case workbook.

The workbook is expected to contain a ``TestCases`` sheet with at least:
case_id, scope_label, user_query, expected_scope_decision.

This runner is intentionally a behavioral smoke test, not an LLM judge. It
checks that the chat layer makes the right high-level decision:

* IN_SCOPE: answer instead of abstaining.
* OUT_OF_SCOPE: abstain/refuse/redirect without citations.
* BORDERLINE: ask for clarification or narrow the request.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.generation.template_answerer import NO_ANSWER_MESSAGE  # noqa: E402
from workflows.online_pipeline import get_local_qa_answer  # noqa: E402


DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "eval" / "chat_test_case_report.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run ChronoRAG chat test cases from XLSX.")
    parser.add_argument("--input", required=True, help="Path to chrono_rag_assistant_1000_test_cases.xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="JSON report output path")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of cases; 0 means all")
    parser.add_argument(
        "--use-current-llm",
        action="store_true",
        help="Use current LLM_PROVIDER. Default forces LLM_PROVIDER=mock for deterministic checks.",
    )
    parser.add_argument(
        "--fail-on-threshold",
        type=float,
        default=0.0,
        help="Exit 1 if pass rate is below this percentage. 0 disables threshold failure.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    if not args.use_current_llm:
        os.environ["LLM_PROVIDER"] = "mock"
        os.environ["CHAT_ROUTER"] = "rules"
        os.environ["CHAT_ANSWERER"] = "raw"

    rows = load_test_cases(input_path)
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    report = run_cases(rows)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    print_summary(report, output_path)
    threshold = float(args.fail_on_threshold or 0)
    if threshold and report["summary"]["pass_rate"] < threshold:
        raise SystemExit(1)


def load_test_cases(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "TestCases" not in wb.sheetnames:
        raise SystemExit("Workbook must contain a TestCases sheet")
    ws = wb["TestCases"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    required = {"case_id", "scope_label", "user_query", "expected_scope_decision"}
    missing = sorted(required - set(headers))
    if missing:
        raise SystemExit(f"TestCases sheet missing required columns: {missing}")

    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        if row.get("case_id") and row.get("user_query"):
            out.append(row)
    return out


def run_cases(rows: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    summary = Counter()
    by_scope: Dict[str, Counter] = defaultdict(Counter)
    by_expected_decision: Dict[str, Counter] = defaultdict(Counter)
    failures: List[Dict[str, Any]] = []
    examples: List[Dict[str, Any]] = []

    for row in rows:
        case_id = str(row.get("case_id") or "").strip()
        scope = str(row.get("scope_label") or "").strip().upper()
        expected_decision = str(row.get("expected_scope_decision") or "").strip().lower()
        query = str(row.get("user_query") or "").strip()

        result = get_local_qa_answer(_topic_from_row(row), query)
        verdict = evaluate_case(scope, expected_decision, result)

        summary["total"] += 1
        summary["pass" if verdict["passed"] else "fail"] += 1
        by_scope[scope]["total"] += 1
        by_scope[scope]["pass" if verdict["passed"] else "fail"] += 1
        by_expected_decision[expected_decision]["total"] += 1
        by_expected_decision[expected_decision]["pass" if verdict["passed"] else "fail"] += 1

        compact = {
            "case_id": case_id,
            "scope_label": scope,
            "expected_scope_decision": expected_decision,
            "query": query,
            "passed": verdict["passed"],
            "reason": verdict["reason"],
            "category": verdict["category"],
            "mode": result.get("mode"),
            "provider": result.get("provider"),
            "model": result.get("model"),
            "citation_count": len(result.get("citations") or []),
            "answer_preview": str(result.get("answer") or "")[:260],
        }
        if not verdict["passed"] and len(failures) < 80:
            failures.append(compact)
        elif len(examples) < 30:
            examples.append(compact)

    total = summary["total"]
    pass_count = summary["pass"]
    return {
        "summary": {
            "total": total,
            "passed": pass_count,
            "failed": summary["fail"],
            "pass_rate": round((pass_count / total * 100) if total else 0.0, 2),
        },
        "by_scope": _counter_map(by_scope),
        "by_expected_scope_decision": _counter_map(by_expected_decision),
        "failures_sample": failures,
        "examples_sample": examples,
    }


def evaluate_case(scope: str, expected_decision: str, result: Dict[str, Any]) -> Dict[str, Any]:
    answer = str(result.get("answer") or "").strip()
    citations = result.get("citations") or []
    category = categorize_response(answer, citations, result)

    if scope == "OUT_OF_SCOPE" or expected_decision == "refuse_or_redirect":
        passed = category in {"abstain", "refusal"}
        return {"passed": passed, "category": category, "reason": "expected_refusal"}

    if scope == "BORDERLINE" or expected_decision == "ask_clarifying_or_narrow_answer":
        passed = category in {"clarification", "project", "meta"} and not citations
        return {"passed": passed, "category": category, "reason": "expected_clarification"}

    # In-scope questions can be corpus-grounded, project guidance, or app meta.
    # They should not silently abstain/refuse.
    passed = category not in {"empty", "abstain", "refusal"}
    return {"passed": passed, "category": category, "reason": "expected_answer"}


def categorize_response(answer: str, citations: List[Dict[str, Any]], result: Dict[str, Any]) -> str:
    if not answer:
        return "empty"
    if answer == NO_ANSWER_MESSAGE:
        return "abstain"
    mode = str(result.get("mode") or "").strip().lower()
    provider = str(result.get("provider") or "").strip().lower()
    if mode in {"project", "llm"} and not citations:
        return "project"
    if provider in {"project-fallback", "scope_guard"} and not citations:
        return "project"
    lowered = answer.lower()
    if citations:
        return "corpus"
    if any(term in lowered for term in ("ngoài phạm vi", "outside scope", "không thể hỗ trợ", "không tìm thấy")):
        return "refusal"
    if any(term in lowered for term in ("hỏi lại", "làm rõ", "clarification", "clarify", "narrow")):
        return "clarification"
    if any(term in lowered for term in ("đại ka có thể hỏi", "chronorag", "metrics hiện tại", "mình là chronorag")):
        return "meta"
    return "project"


def _topic_from_row(row: Dict[str, Any]) -> str:
    domain = str(row.get("domain") or "").strip().lower()
    sub_area = str(row.get("sub_area") or "").strip().lower()
    text = f"{domain} {sub_area}"
    if "distillation" in text or "kd" in text:
        return "knowledge_distillation"
    if "agent" in text:
        return "ai_agent"
    return "rag"


def _counter_map(data: Dict[str, Counter]) -> Dict[str, Dict[str, int]]:
    return {key or "(blank)": dict(counter) for key, counter in sorted(data.items())}


def print_summary(report: Dict[str, Any], output_path: Path) -> None:
    summary = report["summary"]
    print(
        f"Chat test cases: {summary['passed']}/{summary['total']} passed "
        f"({summary['pass_rate']}%). Report: {output_path}"
    )
    for scope, counts in report["by_scope"].items():
        total = counts.get("total", 0)
        passed = counts.get("pass", 0)
        rate = round((passed / total * 100) if total else 0.0, 2)
        print(f"  {scope}: {passed}/{total} ({rate}%)")
    if report["failures_sample"]:
        print("Failure samples:")
        for item in report["failures_sample"][:10]:
            print(f"  {item['case_id']} [{item['scope_label']}]: {item['query']} -> {item['category']}")


if __name__ == "__main__":
    main()
